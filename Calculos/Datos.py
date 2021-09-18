# -*- coding: utf-8 -*-
"""
Created on Fri Mar 15 10:03:15 2019

@author: Samuel
"""
class Datos(): # Clase para leer los datos de climatología directa en otros códigos
    def horario_palma(): #Función para los horarios de Palma
        import openpyxl      #Bibloteca  para trabajar de python con excel  
        import os
        ruta_abs = os.path.dirname(os.path.abspath(__file__)) # Obtención de la ruta absoluta del archivi
        ruta_excel = os.path.join(ruta_abs, 'Data.xlsx')
        doc = openpyxl.load_workbook(ruta_excel)  # Abrir el Excel de Palma
        hoja = doc.get_sheet_by_name('Data') # Leer los datos del Excel
        dias_mes = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31] # Dias que tiene cada mes
        temperaturas_any = ["T_enero", "T_febrero", "T_marzo","T_abril","T_mayo","T_junio","T_julio","T_agosto","T_septiembre","T_octubre","T_noviembre","T_diciembre"]  #Lista de temperaturas de cada mes
        humedad_any = ["HR_enero", "HR_febrero", "HR_marzo","HR_abril","HR_mayo","HR_junio","HR_julio","HR_agosto","HR_septiembre","HR_octubre","HR_noviembre","HR_diciembre"] # Lista de humedades relativas de cada mes
        G_bn_any = ["HR_enero", "HR_febrero", "HR_marzo","HR_abril","HR_mayo","HR_junio","HR_julio","HR_agosto","HR_septiembre","HR_octubre","HR_noviembre","HR_diciembre"] # Lista de humedades relativas de cada mes
        G_d_any = ["HR_enero", "HR_febrero", "HR_marzo","HR_abril","HR_mayo","HR_junio","HR_julio","HR_agosto","HR_septiembre","HR_octubre","HR_noviembre","HR_diciembre"] # Lista de humedades relativas de cada mes        
        v_viento_any = ["HR_enero", "HR_febrero", "HR_marzo","HR_abril","HR_mayo","HR_junio","HR_julio","HR_agosto","HR_septiembre","HR_octubre","HR_noviembre","HR_diciembre"] # Lista de humedades relativas de cada mes        

        count = 17 # Datos empiezan en fila 18
        
        # Generación de listas con las temperaturas de todo el año para generar los diccionarios posteriormente
        for i in range(0,len(temperaturas_any)):
            temperaturas_any[i] = []
            humedad_any[i] = []
            G_bn_any[i] = []
            G_d_any[i] = []
            v_viento_any[i] = []
            for k in range(0,24*dias_mes[i]):
                 count = count + 1                
                 temperaturas_any[i].append(float(hoja.cell(row=count,column = 2).value))
                 humedad_any[i].append(float(hoja.cell(row=count,column = 3).value))
                 G_bn_any[i].append(float(hoja.cell(row=count,column = 5).value))
                 G_d_any[i].append(float(hoja.cell(row=count,column = 6).value))                 
                 v_viento_any[i].append(float(hoja.cell(row=count,column = 8).value))

           
        
        # Generación de los direccionarios que serán las salidas del bloque
        T_ext = {'Enero': temperaturas_any[0], 'Febrero' : temperaturas_any[1], 'Marzo' : temperaturas_any[2], 'Abril' : temperaturas_any[3], 'Mayo' : temperaturas_any[4], 'Junio' : temperaturas_any[5], 'Julio' : temperaturas_any[6], 'Agosto' : temperaturas_any[7], 'Septiembre' : temperaturas_any[8], 'Octubre' : temperaturas_any[9], 'Noviembre' : temperaturas_any[10], 'Diciembre' : temperaturas_any[11]}
        HR_ext = {'Enero': humedad_any[0], 'Febrero' : humedad_any[1], 'Marzo' : humedad_any[2], 'Abril' : humedad_any[3], 'Mayo' : humedad_any[4], 'Junio' : humedad_any[5], 'Julio' : humedad_any[6], 'Agosto' : humedad_any[7], 'Septiembre' : humedad_any[8], 'Octubre' : humedad_any[9], 'Noviembre' : humedad_any[10], 'Diciembre' : humedad_any[11]}    
        G_bn = {'Enero': G_bn_any[0], 'Febrero' : G_bn_any[1], 'Marzo' : G_bn_any[2], 'Abril' : G_bn_any[3], 'Mayo' : G_bn_any[4], 'Junio' : G_bn_any[5], 'Julio' : G_bn_any[6], 'Agosto' : G_bn_any[7], 'Septiembre' : G_bn_any[8], 'Octubre' : G_bn_any[9], 'Noviembre' : G_bn_any[10], 'Diciembre' : G_bn_any[11]}
        G_d = {'Enero': G_d_any[0], 'Febrero' : G_d_any[1], 'Marzo' : G_d_any[2], 'Abril' : G_d_any[3], 'Mayo' : G_d_any[4], 'Junio' : G_d_any[5], 'Julio' : G_d_any[6], 'Agosto' : G_d_any[7], 'Septiembre' : G_d_any[8], 'Octubre' : G_d_any[9], 'Noviembre' : G_d_any[10], 'Diciembre' : G_d_any[11]}
        v_viento = {'Enero': v_viento_any[0], 'Febrero' : v_viento_any[1], 'Marzo' : v_viento_any[2], 'Abril' : v_viento_any[3], 'Mayo' : v_viento_any[4], 'Junio' : v_viento_any[5], 'Julio' : v_viento_any[6], 'Agosto' : v_viento_any[7], 'Septiembre' : v_viento_any[8], 'Octubre' : v_viento_any[9], 'Noviembre' : v_viento_any[10], 'Diciembre' : v_viento_any[11]}
        
        Meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        return T_ext, HR_ext, G_bn, G_d, v_viento, Meses
